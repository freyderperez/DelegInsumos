
"""
DelegInsumos - Servicio de Generación de Reportes
Genera reportes profesionales en PDF y Excel con colores institucionales
"""

import os
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, date, timedelta
from pathlib import Path
from decimal import Decimal
import io

# Librerías para reportes
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.platypus.flowables import Image
from reportlab.lib.colors import HexColor

# Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart
from openpyxl.chart import Reference

# Gráficos
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_agg import FigureCanvasAgg
import pandas as pd

from services.micro_insumos import micro_insumos
from services.micro_empleados import micro_empleados
from services.micro_entregas import micro_entregas
from services.micro_alertas import micro_alertas
from config.config_manager import config
from utils.logger import LoggerMixin, log_operation
from utils.helpers import format_date, safe_filename
from exceptions.custom_exceptions import (
    service_exception_handler,
    ReportGenerationException
)


class ReportColors:
    """Colores institucionales para reportes"""
    AZUL_PRINCIPAL = HexColor("#2196F3")
    AZUL_SECUNDARIO = HexColor("#1976D2") 
    AZUL_CLARO = HexColor("#E3F2FD")
    GRIS = HexColor("#757575")
    GRIS_CLARO = HexColor("#F5F5F5")
    BLANCO = colors.white
    NEGRO = colors.black
    VERDE = HexColor("#4CAF50")
    ROJO = HexColor("#F44336")
    NARANJA = HexColor("#FF9800")


class ReportesService(LoggerMixin):
    """
    Servicio para generación de reportes profesionales en PDF y Excel
    """
    
    def __init__(self):
        super().__init__()
        self.report_config = config.get_reports_config()
        self.output_dir = Path(self.report_config.get('directorio', './reportes/'))
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Configuración de colores institucionales
        self.colors = ReportColors()
        
        self.logger.info("ReportesService inicializado")
    
    def _get_report_filename(self, report_type: str, extension: str = "pdf") -> str:
        """Genera nombre de archivo para reporte"""
        timestamp = datetime.now().strftime(self.report_config.get('formato_fecha_archivo', '%Y%m%d_%H%M%S'))
        safe_type = safe_filename(report_type)
        return f"{safe_type}_{timestamp}.{extension}"
    
    def _create_pdf_header(self, canvas, doc):
        """Crea header estándar para PDFs"""
        canvas.saveState()
        
        # Logo/Título de la empresa
        canvas.setFont("Helvetica-Bold", 16)
        canvas.setFillColor(self.colors.AZUL_PRINCIPAL)
        canvas.drawString(doc.leftMargin, doc.height + doc.bottomMargin - 30, "DelegInsumos")
        
        # Línea decorativa
        canvas.setStrokeColor(self.colors.AZUL_SECUNDARIO)
        canvas.setLineWidth(2)
        canvas.line(doc.leftMargin, doc.height + doc.bottomMargin - 40, 
                   doc.leftMargin + doc.width, doc.height + doc.bottomMargin - 40)
        
        # Fecha de generación
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(self.colors.GRIS)
        canvas.drawRightString(doc.leftMargin + doc.width, doc.height + doc.bottomMargin - 50,
                              f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        canvas.restoreState()
    
    def _create_pdf_footer(self, canvas, doc):
        """Crea footer estándar para PDFs"""
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(self.colors.GRIS)
        
        # Número de página
        page_num = canvas.getPageNumber()
        canvas.drawCentredString(doc.leftMargin + doc.width/2, 30, f"Página {page_num}")
        
        canvas.restoreState()
    
    @service_exception_handler("ReportesService")
    def generar_reporte_inventario_pdf(self, incluir_graficos: bool = True) -> Dict[str, Any]:
        """
        Genera reporte completo de inventario en PDF.
        
        Args:
            incluir_graficos: Si incluir gráficos en el reporte
            
        Returns:
            Diccionario con información del reporte generado
        """
        self.logger.info("Generando reporte de inventario en PDF")
        
        try:
            # Obtener datos
            insumos_data = micro_insumos.listar_insumos(active_only=True, include_status=True)
            alertas_data = micro_alertas.obtener_alertas_dashboard()
            resumen_categorias = micro_insumos.obtener_resumen_por_categoria()
            
            # Crear archivo
            filename = self._get_report_filename("inventario_actual", "pdf")
            filepath = self.output_dir / filename
            
            # Configurar documento
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=A4,
                rightMargin=72, leftMargin=72,
                topMargin=100, bottomMargin=80
            )
            
            # Elementos del reporte
            elements = []
            styles = getSampleStyleSheet()
            
            # Título principal
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=self.colors.AZUL_PRINCIPAL,
                alignment=1
            )
            
            elements.append(Paragraph("Reporte de Inventario Actual", title_style))
            elements.append(Spacer(1, 20))
            
            # Resumen ejecutivo
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=11,
                textColor=self.colors.GRIS
            )
            
            estadisticas = insumos_data.get('statistics', {})
            by_status = insumos_data.get('by_status', {})
            
            summary_text = f"""
            <b>Resumen Ejecutivo:</b><br/>
            • Total de insumos activos: {insumos_data['total']}<br/>
            • Insumos con stock crítico: {by_status.get('criticos', 0)}<br/>
            • Insumos con stock bajo: {by_status.get('bajo_stock', 0)}<br/>
            • Total de alertas activas: {alertas_data['total_active']}<br/>
            """
            
            elements.append(Paragraph(summary_text, summary_style))
            elements.append(Spacer(1, 20))
            
            # Tabla de insumos con alertas
            if by_status.get('criticos', 0) > 0 or by_status.get('bajo_stock', 0) > 0:
                elements.append(Paragraph("Insumos que Requieren Atención", styles['Heading2']))
                
                insumos_alerta = []
                for insumo in insumos_data['insumos']:
                    if (insumo['cantidad_actual'] <= 0 or 
                        insumo['cantidad_actual'] <= insumo['cantidad_minima']):
                        insumos_alerta.append(insumo)
                
                if insumos_alerta:
                    alert_table_data = [['Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Estado']]
                    
                    for insumo in insumos_alerta[:10]:
                        estado = "CRÍTICO" if insumo['cantidad_actual'] <= 0 else "BAJO"
                        alert_table_data.append([
                            insumo['nombre'],
                            insumo['categoria'],
                            f"{insumo['cantidad_actual']} {insumo['unidad_medida']}",
                            f"{insumo['cantidad_minima']} {insumo['unidad_medida']}",
                            estado
                        ])
                    
                    alert_table = Table(alert_table_data, colWidths=[3*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch])
                    alert_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_PRINCIPAL),
                        ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('GRID', (0, 0), (-1, -1), 1, self.colors.AZUL_SECUNDARIO)
                    ]))
                    
                    elements.append(alert_table)
                    elements.append(Spacer(1, 20))
            
            # Secciones adicionales detalladas
            # Distribución por Estado de Stock
            elements.append(Paragraph("Distribución por Estado de Stock", styles['Heading2']))
            estado_table_data = [['Estado', 'Cantidad']]
            estado_table_data.append(['CRÍTICO', str(by_status.get('criticos', 0))])
            estado_table_data.append(['BAJO', str(by_status.get('bajo_stock', 0))])
            estado_table_data.append(['NORMAL', str(by_status.get('normales', 0))])
            estado_table_data.append(['EXCESO', str(by_status.get('exceso', 0))])
            estado_table = Table(estado_table_data, colWidths=[2*inch, 1*inch])
            estado_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_SECUNDARIO),
                ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
            ]))
            elements.append(estado_table)
            elements.append(Spacer(1, 15))

            # Resumen por Categorías
            if resumen_categorias:
                elements.append(Paragraph("Resumen por Categorías", styles['Heading2']))
                cat_table_data = [['Categoría', 'Total Items', 'Cantidad Total', 'Promedio Cantidad']]
                for cat in resumen_categorias[:15]:
                    cat_table_data.append([
                        cat['categoria'],
                        str(cat['total_insumos']),
                        str(cat['cantidad_total']),
                        f"{float(cat['promedio_cantidad']):.2f}"
                    ])
                cat_table = Table(cat_table_data, colWidths=[2.5*inch, 1*inch, 1.2*inch, 1.2*inch])
                cat_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_PRINCIPAL),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                elements.append(cat_table)
                elements.append(Spacer(1, 15))

            # Inventario Detallado (Top 20)
            elements.append(Paragraph("Inventario Detallado (Top 20)", styles['Heading2']))
            # Ordenar por estado de criticidad: crítico, bajo, exceso, normal
            def _estado_rank(item):
                cant = item['cantidad_actual']
                if cant <= 0:
                    return 0
                if cant <= item['cantidad_minima']:
                    return 1
                if cant >= item['cantidad_maxima']:
                    return 2
                return 3
            insumos_ordenados = sorted(insumos_data['insumos'], key=_estado_rank)
            detalle_data = [['Código', 'Nombre', 'Categoría', 'Stock', 'Mín.', 'Máx.', 'Unidad', 'Estado']]
            for ins in insumos_ordenados[:20]:
                if ins['cantidad_actual'] <= 0:
                    estado = "CRÍTICO"
                elif ins['cantidad_actual'] <= ins['cantidad_minima']:
                    estado = "BAJO"
                elif ins['cantidad_actual'] >= ins['cantidad_maxima']:
                    estado = "EXCESO"
                else:
                    estado = "NORMAL"
                detalle_data.append([
                    ins.get('codigo', ins.get('id')),
                    ins['nombre'],
                    ins['categoria'],
                    str(ins['cantidad_actual']),
                    str(ins['cantidad_minima']),
                    str(ins['cantidad_maxima']),
                    ins['unidad_medida'],
                    estado
                ])
            detalle_table = Table(detalle_data, colWidths=[1.2*inch, 2*inch, 1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.9*inch])
            detalle_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_SECUNDARIO),
                ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, self.colors.GRIS)
            ]))
            elements.append(detalle_table)
            elements.append(Spacer(1, 20))

            # Generar PDF
            def add_page_elements(canvas, doc):
                self._create_pdf_header(canvas, doc)
                self._create_pdf_footer(canvas, doc)
            
            doc.build(elements, onFirstPage=add_page_elements, onLaterPages=add_page_elements)
            
            log_operation("REPORTE_PDF_GENERADO", f"Tipo: inventario, Archivo: {filename}")
            self.logger.info(f"Reporte de inventario PDF generado: {filepath}")
            
            return {
                'success': True,
                'filename': filename,
                'filepath': str(filepath),
                'size_mb': round(filepath.stat().st_size / (1024*1024), 2),
                'total_insumos': insumos_data['total'],
                'message': 'Reporte de inventario generado exitosamente en PDF'
            }
            
        except Exception as e:
            self.logger.error(f"Error generando reporte de inventario PDF: {e}")
            raise ReportGenerationException("inventario_pdf", str(e))
    
    @service_exception_handler("ReportesService")
    def generar_reporte_inventario_excel(self) -> Dict[str, Any]:
        """
        Genera reporte de inventario en Excel con formato profesional.
        
        Returns:
            Diccionario con información del reporte generado
        """
        self.logger.info("Generando reporte de inventario en Excel")
        
        try:
            # Obtener datos
            insumos_data = micro_insumos.listar_insumos(active_only=True, include_status=True)
            resumen_categorias = micro_insumos.obtener_resumen_por_categoria()
            alertas_data = micro_alertas.obtener_alertas_dashboard()
            
            # Crear archivo
            filename = self._get_report_filename("inventario_actual", "xlsx")
            filepath = self.output_dir / filename
            
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Hoja 1: Resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            
            # Título
            ws_resumen['A1'] = "Reporte de Inventario - DelegInsumos"
            ws_resumen['A1'].font = Font(name='Calibri', size=16, bold=True, color='1976D2')
            ws_resumen.merge_cells('A1:F1')
            
            # Fecha
            ws_resumen['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws_resumen['A2'].font = Font(name='Calibri', size=10, color='757575')
            
            # Estadísticas
            estadisticas = insumos_data.get('statistics', {})
            by_status = insumos_data.get('by_status', {})
            
            ws_resumen['A4'] = "Estadísticas Generales"
            ws_resumen['A4'].font = Font(name='Calibri', size=12, bold=True, color='2196F3')
            
            stats_data = [
                ['Total de insumos activos:', insumos_data['total']],
                ['Insumos con stock crítico:', by_status.get('criticos', 0)],
                ['Insumos con stock bajo:', by_status.get('bajo_stock', 0)],
                ['Total de alertas activas:', alertas_data['total_active']]
            ]
            
            row = 5
            for label, value in stats_data:
                ws_resumen[f'A{row}'] = label
                ws_resumen[f'B{row}'] = value
                ws_resumen[f'A{row}'].font = Font(name='Calibri', size=10)
                ws_resumen[f'B{row}'].font = Font(name='Calibri', size=10, bold=True)
                row += 1
            
            # Hoja 2: Inventario Completo
            ws_inventario = wb.create_sheet("Inventario Completo")
            
            # Headers
            headers = ['Código', 'Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo',
                      'Stock Máximo', 'Unidad', 'Estado']
            
            for col, header in enumerate(headers, 1):
                cell = ws_inventario.cell(row=1, column=col, value=header)
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de insumos
            for row, insumo in enumerate(insumos_data['insumos'], 2):
                ws_inventario.cell(row=row, column=1, value=insumo.get('codigo', insumo.get('id')))
                ws_inventario.cell(row=row, column=2, value=insumo['nombre'])
                ws_inventario.cell(row=row, column=3, value=insumo['categoria'])
                ws_inventario.cell(row=row, column=4, value=insumo['cantidad_actual'])
                ws_inventario.cell(row=row, column=5, value=insumo['cantidad_minima'])
                ws_inventario.cell(row=row, column=6, value=insumo['cantidad_maxima'])
                ws_inventario.cell(row=row, column=7, value=insumo['unidad_medida'])
                
                # Estado basado en stock
                if insumo['cantidad_actual'] <= 0:
                    estado = "CRÍTICO"
                    color = 'F44336'
                elif insumo['cantidad_actual'] <= insumo['cantidad_minima']:
                    estado = "BAJO"
                    color = 'FF9800'
                elif insumo['cantidad_actual'] >= insumo['cantidad_maxima']:
                    estado = "EXCESO"
                    color = '757575'
                else:
                    estado = "NORMAL"
                    color = '4CAF50'
                
                estado_cell = ws_inventario.cell(row=row, column=8, value=estado)
                estado_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                estado_cell.font = Font(color='FFFFFF', bold=True)
                estado_cell.alignment = Alignment(horizontal='center')
            
            # Autoajustar columnas
            for column in ws_inventario.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_inventario.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            # Hoja 3: Resumen por Categorías
            if resumen_categorias:
                ws_categorias = wb.create_sheet("Por Categorías")
                
                headers_cat = ['Categoría', 'Total Items', 'Cantidad Total', 'Promedio Cantidad']
                
                for col, header in enumerate(headers_cat, 1):
                    cell = ws_categorias.cell(row=1, column=col, value=header)
                    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                for row, categoria in enumerate(resumen_categorias, 2):
                    ws_categorias.cell(row=row, column=1, value=categoria['categoria'])
                    ws_categorias.cell(row=row, column=2, value=categoria['total_insumos'])
                    ws_categorias.cell(row=row, column=3, value=categoria['cantidad_total'])
                    ws_categorias.cell(row=row, column=4, value=float(categoria['promedio_cantidad']))
                
                # Autoajustar columnas
                for column in ws_categorias.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws_categorias.column_dimensions[column_letter].width = max_length + 2
            
            # Guardar archivo
            wb.save(str(filepath))
            
            log_operation("REPORTE_EXCEL_GENERADO", f"Tipo: inventario, Archivo: {filename}")
            self.logger.info(f"Reporte de inventario Excel generado: {filepath}")
            
            return {
                'success': True,
                'filename': filename,
                'filepath': str(filepath),
                'size_mb': round(filepath.stat().st_size / (1024*1024), 2),
                'total_insumos': insumos_data['total'],
                'sheets_created': ['Resumen', 'Inventario Completo', 'Por Categorías'] if resumen_categorias else ['Resumen', 'Inventario Completo'],
                'message': 'Reporte de inventario generado exitosamente en Excel'
            }
            
        except Exception as e:
            self.logger.error(f"Error generando reporte de inventario Excel: {e}")
            raise ReportGenerationException("inventario_excel", str(e))
    
    @service_exception_handler("ReportesService")
    def generar_reporte_entregas_pdf(self, fecha_inicio: Optional[date] = None, 
                                   fecha_fin: Optional[date] = None) -> Dict[str, Any]:
        """
        Genera reporte de entregas en PDF.
        
        Args:
            fecha_inicio: Fecha de inicio del rango (opcional)
            fecha_fin: Fecha de fin del rango (opcional)
            
        Returns:
            Diccionario con información del reporte generado
        """
        self.logger.info("Generando reporte de entregas en PDF")
        
        try:
            # Definir rango de fechas por defecto (últimos 30 días)
            if not fecha_inicio:
                fecha_inicio = date.today() - timedelta(days=30)
            if not fecha_fin:
                fecha_fin = date.today()
            
            # Obtener datos
            entregas_data = micro_entregas.obtener_entregas_por_rango_fechas(fecha_inicio, fecha_fin)
            top_empleados = micro_entregas.obtener_top_empleados_entregas(limit=5)
            top_insumos = micro_entregas.obtener_top_insumos_entregados(limit=5)
            
            # Crear archivo
            filename = self._get_report_filename("entregas_periodo", "pdf")
            filepath = self.output_dir / filename
            
            # Configurar documento
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=A4,
                rightMargin=72, leftMargin=72,
                topMargin=100, bottomMargin=80
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=self.colors.AZUL_PRINCIPAL,
                alignment=1
            )
            
            titulo = f"Reporte de Entregas<br/>{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            elements.append(Paragraph(titulo, title_style))
            elements.append(Spacer(1, 20))
            
            # Resumen del período
            estadisticas = entregas_data.get('statistics', {})
            
            summary_text = f"""
            <b>Resumen del Período:</b><br/>
            • Total de entregas: {entregas_data['total_entregas']}<br/>
            • Empleados únicos: {estadisticas.get('empleados_unicos', 0)}<br/>
            • Insumos únicos: {estadisticas.get('insumos_unicos', 0)}<br/>
            """
            
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=11,
                textColor=self.colors.GRIS
            )
            
            elements.append(Paragraph(summary_text, summary_style))
            elements.append(Spacer(1, 20))
            
            # Top empleados
            if top_empleados:
                elements.append(Paragraph("Top 5 Empleados con Más Entregas", styles['Heading2']))
                
                emp_table_data = [['Empleado', 'Departamento', 'Total Entregas']]
                
                for emp in top_empleados:
                    emp_table_data.append([
                        emp['empleado_nombre'],
                        emp['empleado_departamento'],
                        str(emp['total_entregas'])
                    ])
                
                emp_table = Table(emp_table_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
                emp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_SECUNDARIO),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                
                elements.append(emp_table)
                elements.append(Spacer(1, 15))
            
            # Top insumos
            if top_insumos:
                elements.append(Paragraph("Top 5 Insumos Más Entregados", styles['Heading2']))
                
                ins_table_data = [['Insumo', 'Categoría', 'Cantidad Total', 'Entregas']]
                
                for ins in top_insumos:
                    ins_table_data.append([
                        ins['insumo_nombre'],
                        ins['insumo_categoria'],
                        f"{ins['cantidad_total']} {ins['insumo_unidad']}",
                        str(ins['total_entregas'])
                    ])
                
                ins_table = Table(ins_table_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
                ins_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.VERDE),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                
                elements.append(ins_table)
            
            # Secciones adicionales detalladas del período
            entregas_list = entregas_data.get('entregas', [])
            # Resumen por Empleado
            if entregas_list:
                elements.append(Paragraph("Resumen por Empleado", styles['Heading2']))
                emp_summary = {}
                for e in entregas_list:
                    key = (e.get('empleado_nombre', 'N/A'), e.get('empleado_departamento', ''))
                    info = emp_summary.setdefault(key, {'entregas': 0, 'cantidad_total': 0})
                    info['entregas'] += 1
                    try:
                        info['cantidad_total'] += float(e.get('cantidad', 0))
                    except Exception:
                        pass
                emp_rows = sorted(emp_summary.items(), key=lambda x: (-x[1]['entregas'], -x[1]['cantidad_total']))[:10]
                emp_table_data = [['Empleado', 'Departamento', 'Entregas', 'Cantidad Total']]
                for (nombre, depto), info in emp_rows:
                    emp_table_data.append([nombre, depto, str(info['entregas']), f"{info['cantidad_total']:.0f}"])
                emp_table2 = Table(emp_table_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.2*inch])
                emp_table2.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_PRINCIPAL),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                elements.append(emp_table2)
                elements.append(Spacer(1, 15))

                # Resumen por Insumo
                elements.append(Paragraph("Resumen por Insumo", styles['Heading2']))
                ins_summary = {}
                for e in entregas_list:
                    key = (e.get('insumo_nombre', 'N/A'), e.get('insumo_categoria', ''), e.get('insumo_unidad', ''))
                    info = ins_summary.setdefault(key, {'entregas': 0, 'cantidad_total': 0})
                    info['entregas'] += 1
                    try:
                        info['cantidad_total'] += float(e.get('cantidad', 0))
                    except Exception:
                        pass
                ins_rows = sorted(ins_summary.items(), key=lambda x: (-x[1]['cantidad_total'], -x[1]['entregas']))[:10]
                ins_table_data2 = [['Insumo', 'Categoría', 'Cantidad Total', 'Entregas']]
                for (nombre, cat, unidad), info in ins_rows:
                    ins_table_data2.append([nombre, cat, f"{info['cantidad_total']:.0f} {unidad}", str(info['entregas'])])
                ins_table2 = Table(ins_table_data2, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
                ins_table2.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.VERDE),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                elements.append(ins_table2)
                elements.append(Spacer(1, 15))

                # Entregas por Día
                elements.append(Paragraph("Entregas por Día", styles['Heading2']))
                dia_counts = {}
                for e in entregas_list:
                    try:
                        f = datetime.fromisoformat(e['fecha_entrega'].replace('Z', '+00:00')).strftime('%d/%m/%Y')
                    except Exception:
                        f = e.get('fecha_entrega', '')[:10]
                    info = dia_counts.setdefault(f, {'entregas': 0, 'cantidad_total': 0})
                    info['entregas'] += 1
                    try:
                        info['cantidad_total'] += float(e.get('cantidad', 0))
                    except Exception:
                        pass
                dia_rows = sorted(dia_counts.items(), key=lambda x: x[0], reverse=True)[:10]
                dia_table_data = [['Fecha', 'Entregas', 'Cantidad Total']]
                for f, info in dia_rows:
                    dia_table_data.append([f, str(info['entregas']), f"{info['cantidad_total']:.0f}"])
                dia_table = Table(dia_table_data, colWidths=[1.5*inch, 1*inch, 1.2*inch])
                dia_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.AZUL_SECUNDARIO),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                elements.append(dia_table)
                elements.append(Spacer(1, 15))

                # Detalle de Entregas (Top 20 recientes)
                elements.append(Paragraph("Detalle de Entregas (Top 20 recientes)", styles['Heading2']))
                recientes = sorted(entregas_list, key=lambda e: e.get('fecha_entrega', ''), reverse=True)[:20]
                det_table_data = [['Empleado', 'Insumo', 'Cantidad', 'Unidad', 'Fecha']]
                for e in recientes:
                    try:
                        f = datetime.fromisoformat(e['fecha_entrega'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        f = e.get('fecha_entrega', '')
                    det_table_data.append([
                        e.get('empleado_nombre', 'N/A'),
                        e.get('insumo_nombre', 'N/A'),
                        str(e.get('cantidad', 0)),
                        e.get('insumo_unidad', ''),
                        f
                    ])
                det_table = Table(det_table_data, colWidths=[2.2*inch, 2.2*inch, 0.8*inch, 0.8*inch, 1.0*inch])
                det_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.NARANJA),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, self.colors.GRIS)
                ]))
                elements.append(det_table)
                elements.append(Spacer(1, 20))

            # Generar PDF
            def add_page_elements(canvas, doc):
                self._create_pdf_header(canvas, doc)
                self._create_pdf_footer(canvas, doc)
            
            doc.build(elements, onFirstPage=add_page_elements, onLaterPages=add_page_elements)
            
            log_operation("REPORTE_ENTREGAS_PDF_GENERADO", f"Archivo: {filename}")
            self.logger.info(f"Reporte de entregas PDF generado: {filepath}")
            
            return {
                'success': True,
                'filename': filename,
                'filepath': str(filepath),
                'size_mb': round(filepath.stat().st_size / (1024*1024), 2),
                'total_entregas': entregas_data['total_entregas'],
                'periodo_inicio': fecha_inicio.isoformat(),
                'periodo_fin': fecha_fin.isoformat(),
                'message': f'Reporte de entregas generado exitosamente para el período {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
            }
            
        except Exception as e:
            self.logger.error(f"Error generando reporte de entregas PDF: {e}")
            raise ReportGenerationException("entregas_pdf", str(e))
    
    @service_exception_handler("ReportesService")
    def generar_reporte_alertas_pdf(self) -> Dict[str, Any]:
        """
        Genera reporte de alertas del sistema en PDF.
        
        Returns:
            Diccionario con información del reporte generado
        """
        self.logger.info("Generando reporte de alertas en PDF")
        
        try:
            # Obtener datos de alertas
            alertas_dashboard = micro_alertas.obtener_alertas_dashboard()
            alertas_activas = micro_alertas.obtener_alertas_activas()
            resumen_alertas = micro_alertas.obtener_resumen_alertas()
            
            # Crear archivo
            filename = self._get_report_filename("alertas_sistema", "pdf")
            filepath = self.output_dir / filename
            
            # Configurar documento
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=A4,
                rightMargin=72, leftMargin=72,
                topMargin=100, bottomMargin=80
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=self.colors.ROJO,
                alignment=1
            )
            
            elements.append(Paragraph("Reporte de Alertas del Sistema", title_style))
            elements.append(Spacer(1, 20))
            
            # Resumen de alertas
            summary_text = f"""
            <b>Estado Actual de Alertas:</b><br/>
            • Total alertas activas: {resumen_alertas['total_active']}<br/>
            • Alertas críticas: {resumen_alertas['by_severity'].get('CRITICAL', 0)}<br/>
            • Alertas de alta prioridad: {resumen_alertas['by_severity'].get('HIGH', 0)}<br/>
            • Alertas que requieren acción: {resumen_alertas['action_required']}<br/>
            """
            
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=11,
                textColor=self.colors.GRIS
            )
            
            elements.append(Paragraph(summary_text, summary_style))
            elements.append(Spacer(1, 20))
            
            # Alertas críticas
            if alertas_dashboard['critical']:
                elements.append(Paragraph("Alertas Críticas", styles['Heading2']))
                
                crit_table_data = [['Título', 'Mensaje', 'Fecha', 'Estado']]
                
                for alert in alertas_dashboard['critical']:
                    fecha_alert = datetime.fromisoformat(alert['created_at'].replace('Z', '+00:00'))
                    crit_table_data.append([
                        alert['title'][:30] + '...' if len(alert['title']) > 30 else alert['title'],
                        alert['message'][:50] + '...' if len(alert['message']) > 50 else alert['message'],
                        fecha_alert.strftime('%d/%m/%Y %H:%M'),
                        "Activa" if not alert['resolved'] else "Resuelta"
                    ])
                
                crit_table = Table(crit_table_data, colWidths=[2*inch, 2.5*inch, 1*inch, 0.8*inch])
                crit_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.ROJO),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                
                elements.append(crit_table)
                elements.append(Spacer(1, 20))
            
            # Distribución por tipo
            elements.append(Paragraph("Distribución de Alertas por Tipo", styles['Heading2']))
            
            tipo_table_data = [['Tipo de Alerta', 'Cantidad']]
            
            for tipo, cantidad in resumen_alertas['by_type'].items():
                if cantidad > 0:
                    tipo_display = tipo.replace('_', ' ').title()
                    tipo_table_data.append([tipo_display, str(cantidad)])
            
            if len(tipo_table_data) > 1:
                tipo_table = Table(tipo_table_data, colWidths=[3*inch, 1*inch])
                tipo_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.NARANJA),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
                ]))
                
                elements.append(tipo_table)
            
            # Secciones adicionales de detalle de alertas
            # Resumen por Severidad
            elements.append(Paragraph("Resumen por Severidad", styles['Heading2']))
            sev_table_data = [['Severidad', 'Cantidad']]
            for sev in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                sev_table_data.append([sev.title(), str(resumen_alertas['by_severity'].get(sev, 0))])
            sev_table = Table(sev_table_data, colWidths=[2*inch, 1*inch])
            sev_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors.ROJO),
                ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, self.colors.GRIS)
            ]))
            elements.append(sev_table)
            elements.append(Spacer(1, 15))

            # Listado de Alertas Activas
            if alertas_activas:
                elements.append(Paragraph("Listado de Alertas Activas", styles['Heading2']))
                list_table_data = [['Tipo', 'Título', 'Mensaje', 'Fecha', 'Acción']]
                for a in alertas_activas[:20]:
                    try:
                        f = datetime.fromisoformat(a['created_at'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
                    except Exception:
                        f = a.get('created_at', '')
                    list_table_data.append([
                        a['alert_type'].replace('_', ' ').title(),
                        a['title'][:30] + '...' if len(a['title']) > 30 else a['title'],
                        a['message'][:40] + '...' if len(a['message']) > 40 else a['message'],
                        f,
                        'Sí' if a.get('action_required') else 'No'
                    ])
                list_table = Table(list_table_data, colWidths=[1.4*inch, 2*inch, 2.2*inch, 1.1*inch, 0.6*inch])
                list_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), self.colors.NARANJA),
                    ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.BLANCO),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, self.colors.GRIS)
                ]))
                elements.append(list_table)
                elements.append(Spacer(1, 15))

            # Generar PDF
            def add_page_elements(canvas, doc):
                self._create_pdf_header(canvas, doc)
                self._create_pdf_footer(canvas, doc)
            
            doc.build(elements, onFirstPage=add_page_elements, onLaterPages=add_page_elements)
            
            log_operation("REPORTE_ALERTAS_PDF_GENERADO", f"Archivo: {filename}")
            self.logger.info(f"Reporte de alertas PDF generado: {filepath}")
            
            return {
                'success': True,
                'filename': filename,
                'filepath': str(filepath),
                'size_mb': round(filepath.stat().st_size / (1024*1024), 2),
                'total_alertas': resumen_alertas['total_active'],
                'alertas_criticas': resumen_alertas['by_severity'].get('CRITICAL', 0),
                'message': 'Reporte de alertas generado exitosamente en PDF'
            }
            
        except Exception as e:
            self.logger.error(f"Error generando reporte de alertas PDF: {e}")
            raise ReportGenerationException("alertas_pdf", str(e))
    
    @service_exception_handler("ReportesService")
    def listar_reportes_disponibles(self) -> List[Dict[str, Any]]:
        """
        Lista todos los reportes disponibles en el directorio.
        
        Returns:
            Lista de reportes con información detallada
        """
        self.logger.debug("Listando reportes disponibles")
        
        try:
            reportes = []
            
            # Escanear directorio de reportes
            if self.output_dir.exists():
                for file_path in self.output_dir.iterdir():
                    if file_path.is_file() and file_path.suffix in ['.pdf', '.xlsx']:
                        # Obtener información del archivo
                        stat = file_path.stat()
                        
                        # Determinar tipo de reporte basado en el nombre
                        name_lower = file_path.stem.lower()
                        if 'inventario' in name_lower:
                            report_type = 'Inventario'
                            icon = '📦'
                        elif 'entregas' in name_lower:
                            report_type = 'Entregas'
                            icon = '📋'
                        elif 'empleados' in name_lower:
                            report_type = 'Empleados'
                            icon = '👥'
                        elif 'alertas' in name_lower:
                            report_type = 'Alertas'
                            icon = '⚠️'
                        else:
                            report_type = 'General'
                            icon = '📄'
                        
                        reporte_info = {
                            'filename': file_path.name,
                            'filepath': str(file_path),
                            'type': report_type,
                            'format': file_path.suffix.upper().replace('.', ''),
                            'size_mb': round(stat.st_size / (1024*1024), 2),
                            'size_bytes': stat.st_size,
                            'created_at': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                            'modified_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                            'icon': icon
                        }
                        
                        reportes.append(reporte_info)
            
            # Ordenar por fecha de modificación descendente
            reportes.sort(key=lambda x: x['modified_at'], reverse=True)
            
            return reportes
            
        except Exception as e:
            self.logger.error(f"Error listando reportes: {e}")
            return []
    
    @service_exception_handler("ReportesService")
    def eliminar_reporte(self, filename: str) -> Dict[str, Any]:
        """
        Elimina un reporte específico.
        
        Args:
            filename: Nombre del archivo a eliminar
            
        Returns:
            Resultado de la operación
        """
        try:
            filepath = self.output_dir / filename
            
            if not filepath.exists():
                return {
                    'success': False,
                    'message': f'El archivo {filename} no existe'
                }
            
            # Verificar que es un archivo válido
            if filepath.suffix not in ['.pdf', '.xlsx']:
                return {
                    'success': False,
                    'message': 'Solo se pueden eliminar archivos PDF y Excel'
                }
            
            # Eliminar archivo
            filepath.unlink()
            
            log_operation("REPORTE_ELIMINADO", f"Archivo: {filename}")
            self.logger.info(f"Reporte eliminado: {filename}")
            
            return {
                'success': True,
                'message': f'Reporte {filename} eliminado exitosamente'
            }
            
        except Exception as e:
            self.logger.error(f"Error eliminando reporte {filename}: {e}")
            return {
                'success': False,
                'message': f'Error eliminando reporte: {str(e)}'
            }
    
    @service_exception_handler("ReportesService")
    def limpiar_reportes_antiguos(self, days_old: int = 30) -> Dict[str, Any]:
        """
        Limpia reportes antiguos del directorio.
        
        Args:
            days_old: Días de antigüedad para eliminar
            
        Returns:
            Resultado de la operación
        """
        try:
            if not self.output_dir.exists():
                return {
                    'success': True,
                    'deleted_count': 0,
                    'message': 'Directorio de reportes no existe'
                }
            
            cutoff_date = datetime.now() - timedelta(days=days_old)
            deleted_count = 0
            
            for file_path in self.output_dir.iterdir():
                if (file_path.is_file() and 
                    file_path.suffix in ['.pdf', '.xlsx'] and
                    datetime.fromtimestamp(file_path.stat().st_mtime) < cutoff_date):
                    
                    file_path.unlink()
                    deleted_count += 1
                    self.logger.debug(f"Reporte antiguo eliminado: {file_path.name}")
            
            log_operation("REPORTES_ANTIGUOS_LIMPIADOS", f"Eliminados: {deleted_count}")
            
            return {
                'success': True,
                'deleted_count': deleted_count,
                'cutoff_date': cutoff_date.isoformat(),
                'message': f'Se eliminaron {deleted_count} reportes antiguos'
            }
            
        except Exception as e:
            self.logger.error(f"Error limpiando reportes antiguos: {e}")
            return {
                'success': False,
                'message': f'Error en limpieza: {str(e)}'
            }
    
    @service_exception_handler("ReportesService")
    def obtener_estadisticas_reportes(self) -> Dict[str, Any]:
        """
        Obtiene estadísticas de los reportes generados.
        
        Returns:
            Estadísticas de reportes
        """
        try:
            reportes = self.listar_reportes_disponibles()
            
            # Contar por tipo y formato
            stats = {
                'total_reportes': len(reportes),
                'total_size_mb': sum(r['size_mb'] for r in reportes),
                'by_type': {},
                'by_format': {},
                'recent_reports': [],
                'oldest_report': None,
                'newest_report': None
            }
            
            if reportes:
                # Contar por tipo
                for reporte in reportes:
                    tipo = reporte['type']
                    formato = reporte['format']
                    
                    stats['by_type'][tipo] = stats['by_type'].get(tipo, 0) + 1
                    stats['by_format'][formato] = stats['by_format'].get(formato, 0) + 1
                
                # Reportes recientes (últimos 5)
                stats['recent_reports'] = reportes[:5]
                
                # Más antiguo y más nuevo
                reportes_por_fecha = sorted(reportes, key=lambda x: x['created_at'])
                stats['oldest_report'] = reportes_por_fecha[0] if reportes_por_fecha else None
                stats['newest_report'] = reportes_por_fecha[-1] if reportes_por_fecha else None
            
            return stats
            
        except Exception as e:
            self.logger.error(f"Error obteniendo estadísticas de reportes: {e}")
            return {
                'total_reportes': 0,
                'error': str(e)
            }


# Instancia global del servicio de reportes
reportes_service = ReportesService()

# Funciones de conveniencia para uso directo
def generar_reporte_inventario_pdf(incluir_graficos: bool = True) -> Dict[str, Any]:
    """Función de conveniencia para generar reporte de inventario PDF"""
    return reportes_service.generar_reporte_inventario_pdf(incluir_graficos)

def generar_reporte_inventario_excel() -> Dict[str, Any]:
    """Función de conveniencia para generar reporte de inventario Excel"""
    return reportes_service.generar_reporte_inventario_excel()

def generar_reporte_entregas_pdf(fecha_inicio: Optional[date] = None, 
                               fecha_fin: Optional[date] = None) -> Dict[str, Any]:
    """Función de conveniencia para generar reporte de entregas PDF"""
    return reportes_service.generar_reporte_entregas_pdf(fecha_inicio, fecha_fin)

def generar_reporte_alertas_pdf() -> Dict[str, Any]:
    """Función de conveniencia para generar reporte de alertas PDF"""
    return reportes_service.generar_reporte_alertas_pdf()

def listar_reportes_disponibles() -> List[Dict[str, Any]]:
    """Función de conveniencia para listar reportes disponibles"""
    return reportes_service.listar_reportes_disponibles()